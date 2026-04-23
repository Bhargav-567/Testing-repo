import pandas as pd
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

def generate_exam_dataframe():
    """Returns the sample exam DataFrame"""

    df = pd.DataFrame({
        'Q_ID': ['Q1', 'Q2', 'Q3', 'Q4', 'Q5', 'Q6'],

        'Question': [
            'What is function of boiler drum?',
            'Explain steam turbine function',
            'What is economiser purpose?',
            'Superheater function?',
            'Turbine converts energy:',
            'Boiler drum separates:'
        ],

        'Type': ['mcq', 'descriptive', 'descriptive', 'descriptive', 'mcq', 'mcq'],

        'option1': [
            'Steam generator', '', '', '',
            'Thermal→thermal', 'Steam from air'
        ],
        'option2': [
            'Water separator', '', '', '',
            'Thermal→mechanical', 'Steam from water'
        ],
        'option3': [
            'Heat exchanger', '', '', '',
            'Mechanical→electrical', 'Water from steam'
        ],
        'option4': [
            'Power source', '', '', '',
            'Chemical→electrical', 'Air from water'
        ],

        'Teacher_Answer': [
            'B',
            'Steam turbine drives generator to produce electricity',
            'Economiser preheats feedwater using flue gas heat',
            'Superheater raises steam temperature above saturation',
            'C',
            'B'
        ],

        'Max_Score': [1, 10, 8, 10, 1, 1],

        'Concept_Names': [
            'boiler_drum',
            'steam_turbine,generator',
            'economiser,feedwater',
            'superheater,saturation',
            'turbine,energy_conversion',
            'drum,steam_separation'
        ],

        'Concept_Keywords': [
            'drum,boiler,separate,steam,water',
            'steam,turbine,blades,rotate;generator,shaft,electricity,power',
            'economiser,preheat,flue,gas;feedwater,boiler,inlet,temperature',
            'superheater,steam,heat,furnace;saturation,dry,enthalpy,efficiency',
            'turbine,convert,mechanical,electrical;energy,thermal,mechanical',
            'drum,separate,cylinder,boiler;steam,saturated,vapor,water'
        ]
    })

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="Exam_Questions_Complete_Sample.xlsx"'
    )

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Descriptive_Master', index=False)
        ws = writer.sheets['Descriptive_Master']

        # Auto-size columns (robust + readable)
        for idx, column in enumerate(ws.columns, 1):
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    return response


def download_sample_excel(request):
    df = generate_exam_dataframe()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        'attachment; filename="Exam_Questions_Complete_Sample.xlsx"'
    )

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Descriptive_Master', index=False)
        ws = writer.sheets['Descriptive_Master']

        # Auto-size columns (robust + readable)
        for idx, column in enumerate(ws.columns, 1):
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column
            )
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    return response

def download_sample_csv(request):
    df = generate_exam_dataframe()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        'attachment; filename="Exam_Questions_Complete_Sample.csv"'
    )

    df.to_csv(response, index=False)
    

generate_exam_dataframe()